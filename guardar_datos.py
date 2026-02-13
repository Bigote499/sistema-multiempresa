import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from db_manager import insertar_empresa
import sqlite3

def limpiar_texto(t):
    """Limpia texto para evitar errores de codificación y espacios raros."""
    return (
        str(t)
        .encode("utf-8", errors="ignore")
        .decode("utf-8")
        .strip()
    )

def validar_cuit(cuit):
    """Valida que el CUIT tenga 11 dígitos numéricos."""
    cuit = str(cuit).strip()
    numeros = ''.join(filter(str.isdigit, cuit))
    if len(numeros) == 11:
        return f"{numeros[0:2]}-{numeros[2:10]}-{numeros[10]}"
    return None

def normalizar_columnas(df):
    """
    Normaliza nombres de columnas a los esperados por la base y exportaciones.
    Devuelve una copia del DataFrame con columnas estandarizadas y orden fijo.
    """
    mapa = {
        "Nombre": "nombre",
        "CUIT": "cuit",
        "cuit": "cuit",
        "Email": "email",
        "web": "web",
        "Domicilio": "domicilio",
        "Contacto": "contacto",
    }
    df2 = df.copy()
    df2 = df2.rename(columns={k: v for k, v in mapa.items() if k in df2.columns})
    # Asegurar tipos string y limpieza básica
    for col in df2.columns:
        df2[col] = df2[col].astype(str).fillna("").map(limpiar_texto)
    # Orden fijo de columnas
    orden = ["nombre", "cuit", "email", "web", "domicilio", "contacto"]
    df2 = df2[[c for c in orden if c in df2.columns]]
    return df2

def guardar_en_db(cliente, df):
    resumen = {"guardadas": 0, "omitidas": 0, "errores": 0, "cuit_invalidos": 0}

    try:
        df_norm = normalizar_columnas(df)

        def es_valido(valor):
            v = str(valor).strip().lower()
            return v and v != "nan"

        for idx, fila in df_norm.iterrows():
            try:
                nombre = limpiar_texto(fila.get("nombre", ""))
                cuit_raw = limpiar_texto(fila.get("cuit", ""))
                cuit_formateado = validar_cuit(cuit_raw)
                if cuit_raw and cuit_formateado is None:
                    print(f"[VALIDACIÓN] Fila {idx+1} omitida por CUIT inválido: '{cuit_raw}'")
                    resumen["omitidas"] += 1
                    resumen["cuit_invalidos"] += 1
                    continue
                email = limpiar_texto(fila.get("email", ""))
                web = limpiar_texto(fila.get("web", ""))
                domicilio = limpiar_texto(fila.get("domicilio", ""))
                contacto = limpiar_texto(fila.get("contacto", ""))

                
                # Validación de campos clave
                if not es_valido(nombre) or not es_valido(web):
                    print(f"[VALIDACIÓN] Fila {idx+1} omitida por falta de campos clave: nombre='{nombre}', web='{web}'")
                    resumen["omitidas"] += 1
                    continue

                resultado = insertar_empresa(cliente, nombre, cuit_formateado, email, web, domicilio, contacto)
                if resultado.get("ok"):
                    resumen["guardadas"] += 1
                else:
                    print(f"[ERROR] Fila {idx+1} no se guardó ({resultado.get('motivo')}): {nombre}")
                    resumen["omitidas"] += 1

            except Exception as e_fila:
                print(f"[ERROR] Fila {idx+1} falló inesperadamente: {e_fila}")
                resumen["errores"] += 1

    except Exception as e_general:
        print(f"[ERROR] Fallo general al guardar en DB: {e_general}")
        resumen["errores"] += 1

    return resumen

def db_a_excel(cliente, nombre_archivo="empresas_exportado.xlsx"):
    """Exporta los datos de la base SQLite del cliente a un archivo Excel con formato ajustado."""
    try:
        conn = sqlite3.connect(f"{cliente}.db")
        df = pd.read_sql_query("SELECT * FROM empresas", conn)
        conn.close()

        # Exportar con motor seguro
        with pd.ExcelWriter(nombre_archivo, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="Empresas")

        # Reabrir con openpyxl para ajustar formato
        from openpyxl import load_workbook
        wb = load_workbook(nombre_archivo)
        hoja = wb["Empresas"]

        for i, col in enumerate(df.columns, 1):
            max_largo = max([len(str(valor)) for valor in df[col]] + [len(col)])
            hoja.column_dimensions[get_column_letter(i)].width = max_largo + 2

            for fila in range(2, len(df) + 2):
                celda = hoja.cell(row=fila, column=i)
                celda.alignment = Alignment(wrap_text=True, vertical="top")

        wb.save(nombre_archivo)
        print(f"Excel exportado desde {cliente}.db a {nombre_archivo}")
        return True

    except Exception as e:
        print(f"Error al exportar DB a Excel: {e}")
        return False